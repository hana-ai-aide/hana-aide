#!/usr/bin/env python
# xlsx-convert.py — Excel(.xlsx) → Markdown 的「機械轉換」（忠實、可重複、可排程）。
# SPEC-excel-import §3.1。走與 docx/pptx-convert 完全相同的插槽：--import <檔> <doc_dir> [--source <rel>]。
# 做：
#   ① 多頁籤 → 多段：每個 worksheet 一段 `## <頁籤名>` ＋ 一張 GFM 表格（第一列當表頭）。
#   ② 公式兩讀（openpyxl）：data_only=True 取 Excel 上次存檔算好的「快取值」（顯示用）；
#      data_only=False 取「公式原文」（如 =SUM(A1:A9)，備查用）。儲存格若是公式 → 表格填快取值，
#      並在表格後附一段「公式對照」HTML 註解（儲存格→公式原文），可 grep、可 diff、不擾閱讀。
#   ③ 快取值缺失（該檔從未在 Excel 存過）→ data_only 讀到 None → 顯示公式原文＋「（未計算）」，不自行重算。
#   ④ 只轉「有內容範圍」（避免百萬空列灌進 md），超過上限標示截斷（不靜默吞）。
#   ⑤ 合併儲存格取左上值、其餘留空（Phase 1 不還原跨欄合併版面，屬預覽層）。
#   ⑥ 內嵌圖片（少見）best-effort 抽到 assets/，md 以 ![]() 引用。
# 不做（Phase 2）：公式重算、線上編輯、寫回 source.xlsx。
# 相依：openpyxl（pip install openpyxl；純 Python、跨平台）。
# 用法：python xlsx-convert.py --import "<檔>.xlsx" "<輸出目錄>" [--source "<rel>"]
import os, sys, datetime

MAX_ROWS = 5000   # 單頁籤資料列上限（超過標示截斷，不靜默吞）
MAX_COLS = 200    # 單頁籤欄上限


def _fmt(v):
    """把 openpyxl 儲存格值字串化：日期→ISO、整數值 float→去 .0、None→空字串。"""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    return str(v)


def _esc_cell(s):
    """GFM 表格內防呆：| → \\|、換行 → <br>（否則會拆掉表格結構）。"""
    return s.replace('\\', '\\\\').replace('|', '\\|').replace('\r\n', '<br>').replace('\n', '<br>').replace('\r', '<br>')


def _extract_images(ws, assets_dir, prefix):
    """best-effort 抽出 sheet 內嵌圖片到 assets/，回傳 md 引用清單（多數資料表無圖）。"""
    refs = []
    imgs = getattr(ws, '_images', None) or []
    for i, img in enumerate(imgs, start=1):
        try:
            data = None
            ref = getattr(img, 'ref', None)
            if hasattr(img, '_data') and callable(img._data):
                data = img._data()
            elif hasattr(ref, 'read'):
                data = ref.read()
            if not data:
                continue
            ext = (getattr(img, 'format', None) or 'png').lower()
            if ext == 'jpeg':
                ext = 'jpg'
            os.makedirs(assets_dir, exist_ok=True)
            fname = '%s_img%d.%s' % (prefix, i, ext)
            with open(os.path.join(assets_dir, fname), 'wb') as f:
                f.write(data)
            refs.append('![](assets/%s)' % fname)
        except Exception:
            continue
    return refs


def xlsx_import(xlsx_path, doc_dir, source_rel='', values_overlay=None):
    # values_overlay（XLSX-07 用）：{sheet: {coord: displayValue}} — 回寫後 openpyxl 不保留公式快取，
    # 用編輯器（Univer）重算值補公式格的顯示，讓 doc.md 文字真身在回寫後仍可查找到正確結果。
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    values_overlay = values_overlay or {}

    os.makedirs(doc_dir, exist_ok=True)
    assets_dir = os.path.join(doc_dir, 'assets')

    # 一次開兩份：values=快取值（顯示）、formulas=公式原文（備查）。
    wb_vals = load_workbook(xlsx_path, data_only=True, read_only=False)
    wb_forms = load_workbook(xlsx_path, data_only=False, read_only=False)

    out = []
    for ws_v in wb_vals.worksheets:
        title = ws_v.title
        ws_f = wb_forms[title] if title in wb_forms.sheetnames else None
        out.append('## %s' % title)
        out.append('')

        max_row = ws_v.max_row or 0
        max_col = ws_v.max_column or 0
        # 完全空白頁籤（openpyxl 對空 sheet 常回 max_row=1/max_col=1 但值皆 None）
        if max_row < 1 or max_col < 1:
            out.append('（空白頁籤）')
            out.append('')
            continue

        row_truncated = max_row > MAX_ROWS
        col_truncated = max_col > MAX_COLS
        rows_n = min(max_row, MAX_ROWS)
        cols_n = min(max_col, MAX_COLS)

        # 掃格：填顯示值、收集公式對照。空掉的頁籤在此判定（全 None）。
        formula_refs = []   # (coord, formula原文)
        grid = []           # rows_n × cols_n 的顯示字串
        any_content = False
        for r in range(1, rows_n + 1):
            row_out = []
            for c in range(1, cols_n + 1):
                cached = ws_v.cell(row=r, column=c).value
                formula = ws_f.cell(row=r, column=c).value if ws_f is not None else None
                is_formula = isinstance(formula, str) and formula.startswith('=')
                if is_formula:
                    coord = '%s%d' % (get_column_letter(c), r)
                    formula_refs.append((coord, formula))
                    if cached is None:
                        ov = (values_overlay.get(title) or {}).get(coord)
                        disp = _fmt(ov) if ov is not None else ('%s（未計算）' % formula)
                    else:
                        disp = _fmt(cached)
                else:
                    disp = _fmt(cached)
                if disp != '':
                    any_content = True
                row_out.append(disp)
            grid.append(row_out)

        if not any_content and not formula_refs:
            out.append('（空白頁籤）')
            out.append('')
            continue

        # GFM 表格：第一列當表頭（Excel 資料表慣例）；表頭全空時給欄位字母補位以維持合法表格。
        header = grid[0]
        if all(h == '' for h in header):
            header = [get_column_letter(c) for c in range(1, cols_n + 1)]
            body_rows = grid
        else:
            body_rows = grid[1:]

        out.append('| ' + ' | '.join(_esc_cell(h) for h in header) + ' |')
        out.append('|' + '|'.join(['---'] * cols_n) + '|')
        for row_out in body_rows:
            out.append('| ' + ' | '.join(_esc_cell(x) for x in row_out) + ' |')
        out.append('')

        if row_truncated or col_truncated:
            note = []
            if row_truncated:
                note.append('列 %d→%d' % (max_row, MAX_ROWS))
            if col_truncated:
                note.append('欄 %d→%d' % (max_col, MAX_COLS))
            out.append('> ⚠️ 內容超過上限已截斷（%s）；完整資料請開 source.xlsx。' % '、'.join(note))
            out.append('')

        # 內嵌圖片（少見）
        img_refs = _extract_images(ws_v, assets_dir, _safe_prefix(title))
        for ref in img_refs:
            out.append(ref)
        if img_refs:
            out.append('')

        # 公式對照：HTML 註解（渲染時不干擾閱讀，但仍是純文字 → Hana grep 得到、版本歷史 diff 得到）。
        if formula_refs:
            out.append('<!-- 公式對照（%s）' % title)
            for coord, formula in formula_refs:
                out.append('%s = %s' % (coord, formula[1:]))   # 去掉開頭 '='
            out.append('-->')
            out.append('')

    md_path = os.path.join(doc_dir, 'doc.md')
    header_lines = []
    if source_rel:
        header_lines.append('<!-- source: %s -->' % source_rel)
        header_lines.append('')
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(header_lines + out).rstrip() + '\n')

    print('MD_FILE: ' + md_path)
    print('ASSETS_DIR: ' + assets_dir)


def _safe_prefix(title):
    return ''.join(ch if ch.isalnum() else '_' for ch in title)[:40] or 'sheet'


if __name__ == '__main__':
    if '--import' in sys.argv:
        args = sys.argv[1:]
        ii = args.index('--import')
        rem = args[ii + 1:]
        if len(rem) < 2:
            print('Usage: python xlsx-convert.py --import "<xlsx>" "<doc_dir>" [--source "<rel>"]')
            sys.exit(1)
        xlsx_p = rem[0]
        doc_d = rem[1]
        src_rel = ''
        if '--source' in rem:
            si = rem.index('--source')
            src_rel = rem[si + 1] if si + 1 < len(rem) else ''
        overlay = None
        if '--values' in rem:
            vi = rem.index('--values')
            vpath = rem[vi + 1] if vi + 1 < len(rem) else ''
            if vpath and os.path.exists(vpath):
                import json as _json
                with open(vpath, encoding='utf-8') as _vf:
                    overlay = _json.load(_vf)
        xlsx_import(xlsx_p, doc_d, src_rel, overlay)
    else:
        print('Usage: python xlsx-convert.py --import "<xlsx>" "<doc_dir>" [--source "<rel>"]')
        sys.exit(1)
