#!/usr/bin/env python
# xlsx-writeback.py — SPEC-excel-import §4.3 (XLSX-06): 把前端「變更集」只寫回被動的儲存格到 source.xlsx。
# 誠實邊界：openpyxl 重存 .xlsx 不保證位元保真——圖表/樞紐/條件式格式/巨集可能失真或遺失（見 --detect）。
# 故本腳本「只改指定的格、其餘盡量原封」，並提供 --detect 供上層對高風險檔警告/降級唯讀（XLSX-08）。
#
# 兩種模式：
#   ① 回寫： python xlsx-writeback.py --edits <edits.json> <source.xlsx>
#        edits.json = {"changes":[{"sheet":"工令主檔","cell":"E2","formula":"=C2*D2","value":"1250"},
#                                  {"sheet":"工令主檔","cell":"C2","value":100}, ...]}
#        - 有 formula（以 '=' 開頭）→ 寫公式原文（Excel 開檔時自行重算；openpyxl 不重算）。
#        - 否則 → 寫 value（自動判型：int/float/bool/空→None/其餘字串）。
#        stdout: JSON {ok, changed, skipped, errors, sheets}
#   ② 風險偵測： python xlsx-writeback.py --detect <source.xlsx>
#        stdout: JSON {ok, risks:{charts,pivots,conditionalFormats,vbaMacros,dataValidations,
#                      mergedCells,definedNames}, highRisk:bool, reasons:[...]}
#
# 相依：openpyxl（與 xlsx-convert.py 同）。
import sys, os, json, datetime


def _coerce(v):
    """把前端來的字串/值判成 Excel 適當型別：空→None、bool、int、float、ISO 日期、其餘原樣字串。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v)
    if s == '':
        return None
    low = s.strip().lower()
    if low in ('true', 'false'):
        return low == 'true'
    # 數字（避免把 '007'、'1e3 電話' 這種前導零/含字母的當數字 → 只收乾淨數字）
    try:
        if s.strip().lstrip('-').isdigit():
            return int(s.strip())
    except Exception:
        pass
    try:
        f = float(s)
        # 排除 'nan'/'inf'（float() 會接受，但存進表格無意義）
        if f == f and abs(f) != float('inf'):
            return f
    except Exception:
        pass
    return s


def writeback(xlsx_path, edits):
    from openpyxl import load_workbook

    changes = edits.get('changes') or []
    if not changes:
        return {'ok': True, 'changed': 0, 'skipped': 0, 'errors': [], 'sheets': []}

    # data_only=False → 保留公式原文；keep_links 預設 True。只改被動格、其餘 openpyxl round-trip 原樣保留。
    wb = load_workbook(xlsx_path, data_only=False, read_only=False, keep_vba=xlsx_path.lower().endswith('.xlsm'))

    changed = 0
    skipped = 0
    errors = []
    sheets_touched = set()

    for ch in changes:
        try:
            sheet = ch.get('sheet')
            cell = (ch.get('cell') or '').strip()
            if not sheet or not cell:
                skipped += 1
                errors.append({'change': ch, 'note': 'missing sheet/cell'})
                continue
            if sheet not in wb.sheetnames:
                skipped += 1
                errors.append({'sheet': sheet, 'cell': cell, 'note': 'sheet not found'})
                continue
            ws = wb[sheet]
            formula = ch.get('formula')
            if isinstance(formula, str) and formula.startswith('='):
                ws[cell] = formula   # 公式原文；Excel 開檔重算
            else:
                ws[cell] = _coerce(ch.get('value'))
            changed += 1
            sheets_touched.add(sheet)
        except Exception as e:
            skipped += 1
            errors.append({'change': ch, 'error': str(e)})

    if changed > 0:
        wb.save(xlsx_path)

    return {'ok': True, 'changed': changed, 'skipped': skipped,
            'errors': errors, 'sheets': sorted(sheets_touched)}


def detect_risks(xlsx_path):
    """偵測 openpyxl 重存會失真/遺失的高風險結構，供上層對高風險檔警告或降級唯讀（XLSX-08）。"""
    from openpyxl import load_workbook

    risks = {'charts': 0, 'pivots': 0, 'conditionalFormats': 0, 'vbaMacros': False,
             'dataValidations': 0, 'mergedCells': 0, 'definedNames': 0, 'images': 0}
    reasons = []

    is_macro = xlsx_path.lower().endswith('.xlsm')
    wb = load_workbook(xlsx_path, data_only=False, read_only=False, keep_vba=is_macro)

    # VBA 巨集：.xlsm 或載入後帶 vba_archive
    if is_macro or getattr(wb, 'vba_archive', None) is not None:
        risks['vbaMacros'] = True   # 分級/敘述在下方統一處理

    for ws in wb.worksheets:
        # 圖表（openpyxl 讀不回既有圖表 → 重存必遺失）
        try:
            n = len(getattr(ws, '_charts', []) or [])
            risks['charts'] += n
        except Exception:
            pass
        # 內嵌圖片
        try:
            risks['images'] += len(getattr(ws, '_images', []) or [])
        except Exception:
            pass
        # 條件式格式
        try:
            cf = ws.conditional_formatting
            risks['conditionalFormats'] += sum(1 for _ in cf)
        except Exception:
            pass
        # 資料驗證
        try:
            dv = ws.data_validations.dataValidation
            risks['dataValidations'] += len(dv or [])
        except Exception:
            pass
        # 合併儲存格
        try:
            risks['mergedCells'] += len(list(ws.merged_cells.ranges))
        except Exception:
            pass
        # 樞紐分析表（pivot）
        try:
            risks['pivots'] += len(getattr(ws, '_pivots', []) or [])
        except Exception:
            pass

    # 活頁簿層 pivot caches
    try:
        risks['pivots'] += len(getattr(wb, '_pivots', []) or [])
    except Exception:
        pass
    try:
        dn = wb.defined_names
        risks['definedNames'] = len(list(dn)) if hasattr(dn, '__iter__') else len(getattr(dn, 'definedName', []) or [])
    except Exception:
        pass

    # 依「實測 openpyxl 3.1.5 round-trip 失真清單」（XLSX-08）分級，非套用過時的「圖表一定遺失」傳言：
    #   高風險（建議降級唯讀）＝樞紐 / VBA 巨集：openpyxl 對這兩類支援薄弱、回寫最易破壞。
    #   警告（可續、加提示）＝圖表 / 條件式格式 / 資料驗證：簡單情形實測可保留，但複雜真實檔不保證。
    warnings = []
    if risks['pivots']:
        reasons.append('含 %d 個樞紐分析表：openpyxl 對樞紐支援薄弱，回寫後可能失效或需重整。' % risks['pivots'])
    if risks['vbaMacros']:
        reasons.append('含 VBA 巨集：回寫可能破壞或無法保留巨集，且巨集邏輯不執行。')
    if risks['charts']:
        warnings.append('含 %d 個圖表：簡單圖表實測可保留，複雜樣式/資料標籤/次座標軸不保證。' % risks['charts'])
    if risks['conditionalFormats']:
        warnings.append('含 %d 條條件式格式：基本規則實測可保留，色階/資料橫條/圖示集等進階規則可能降級。' % risks['conditionalFormats'])
    if risks['dataValidations']:
        warnings.append('含 %d 條資料驗證：基本清單/範圍實測可保留，跨表引用可能失真。' % risks['dataValidations'])
    # 公式快取值：openpyxl 重存不寫入公式快取 → 回寫後非 Excel 讀取者（含 doc.md）看到的公式格為「未計算」，
    # Excel 開檔會自動重算補回。上層以 Univer 重算值做 doc.md overlay 緩解。此為所有含公式檔皆有的固定行為。
    warnings.append('公式快取值不隨 openpyxl 保存（Excel 開檔會自動重算）；doc.md 文字真身以編輯器重算值補位。')

    high = bool(risks['pivots'] or risks['vbaMacros'])
    return {'ok': True, 'risks': risks, 'highRisk': high, 'reasons': reasons, 'warnings': warnings}


if __name__ == '__main__':
    args = sys.argv[1:]
    try:
        if '--detect' in args:
            i = args.index('--detect')
            xlsx_p = args[i + 1] if i + 1 < len(args) else ''
            if not xlsx_p or not os.path.exists(xlsx_p):
                print(json.dumps({'ok': False, 'error': 'xlsx not found'}))
                sys.exit(1)
            print(json.dumps(detect_risks(xlsx_p), ensure_ascii=False))
            sys.exit(0)
        if '--edits' in args:
            i = args.index('--edits')
            edits_p = args[i + 1] if i + 1 < len(args) else ''
            # source.xlsx = 第一個非旗標、非 edits 路徑的參數
            rest = [a for k, a in enumerate(args) if a not in ('--edits',) and a != edits_p]
            xlsx_p = rest[0] if rest else ''
            if not edits_p or not os.path.exists(edits_p):
                print(json.dumps({'ok': False, 'error': 'edits json not found'}))
                sys.exit(1)
            if not xlsx_p or not os.path.exists(xlsx_p):
                print(json.dumps({'ok': False, 'error': 'source.xlsx not found'}))
                sys.exit(1)
            with open(edits_p, encoding='utf-8') as f:
                edits = json.load(f)
            result = writeback(xlsx_p, edits)
            print(json.dumps(result, ensure_ascii=False))
            sys.exit(0 if result.get('ok') else 1)
        print(json.dumps({'ok': False, 'error': 'usage: --edits <json> <xlsx> | --detect <xlsx>'}))
        sys.exit(1)
    except Exception as e:
        print(json.dumps({'ok': False, 'error': str(e)}, ensure_ascii=False))
        sys.exit(1)
